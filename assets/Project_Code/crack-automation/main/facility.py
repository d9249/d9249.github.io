import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border,Side
from openpyxl.styles import Alignment, PatternFill
from PIL import Image as IMG
from .models import Category,Crack,CrackObj

thin_border = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
  
top_bold_border = Border(top=Side(style='medium'))
left_bold_border = Border(left=Side(style='medium'))
right_bold_border = Border(right=Side(style='medium'))
bottom_bold_border = Border(bottom=Side(style='medium'))
  
top_left_bold_border = Border(top=Side(style='medium'), left=Side(style='medium'))
top_right_bold_border = Border(top=Side(style='medium'), right=Side(style='medium'))
bottom_left_bold_border = Border(bottom=Side(style='medium'), left=Side(style='medium'))
bottom_right_bold_border = Border(bottom=Side(style='medium'), right=Side(style='medium'))

def facility(wb,pk):
  baseWidth = 420
  baseHeight = 260
  category = Category.objects.get(pk=pk)
  
  grayFill = PatternFill(start_color='CCCCCC',
                        end_color='CCCCCC', fill_type='solid')
  sheet = wb.worksheets[0]
  sheet.title = '시설물 현황'
  sheet.column_dimensions["A"].width = 1

  sheet = wb['시설물 현황']
  sheet['B2'] = "□ 시설물 현황"
  sheet['B3'] = "가. 일반현황"
  sheet['B4'] = '시설물명'
  sheet['B5'] = '시설물위치'
  sheet['B6'] = '용도'
  sheet['B7'] = '구조형식'
  sheet['B8'] = '종별'
  sheet['B9'] = '규모 및 제원 추가사항'

  sheet['E4'] = "시설물번호"
  sheet['E5'] = "준공일자"
  sheet['E6'] = "시설물규모"
  sheet['E7'] = "부대시설"

  sheet['D8'] = '전차안전등급'
  sheet['F8'] = '안전등급결과'

  sheet['B4'].fill = grayFill
  sheet['B5'].fill = grayFill
  sheet['B6'].fill = grayFill
  sheet['B7'].fill = grayFill
  sheet['B9'].fill = grayFill

  sheet['E4'].fill = grayFill
  sheet['E5'].fill = grayFill
  sheet['E6'].fill = grayFill
  sheet['E7'].fill = grayFill

  sheet['B8'].fill = grayFill
  sheet['D8'].fill = grayFill
  sheet['F8'].fill = grayFill

  sheet.merge_cells('C5:D5')
  sheet.merge_cells('C4:D4')
  sheet.merge_cells('C6:D6')
  sheet.merge_cells('C7:D7')

  sheet.merge_cells('F4:G4')
  sheet.merge_cells('F5:G5')
  sheet.merge_cells('F6:G6')
  sheet.merge_cells('F7:G7')

  
  sheet.merge_cells('F4:G4')
  sheet.merge_cells('F5:G5')
  sheet.merge_cells('F6:G6')
  sheet.merge_cells('F7:G7')

  sheet.merge_cells('B9:G9')
  sheet.merge_cells('B10:G10')

  sheet['C4'] = category.facilityName
  sheet['C5'] = category.facilityNo
  sheet['C6'] = category.usage
  sheet['C7'] = category.structuralForm

  sheet['F4'] = category.facilityNo
  sheet['F5'] = category.completionDate
  sheet['F6'] = category.facilityStructure
  sheet['F7'] = category.amenities


  sheet['C8'] = category.floors
  sheet['E8'] = category.grade
  sheet['G8'] = category.testResults

  sheet['B10'] = category.plus

  # cell 테두리
  for row in sheet['B4':'G10']:
      for cell in row:
          cell.border = thin_border

  sheet['B13'].border = top_left_bold_border
  sheet['C13'].border = top_bold_border
  sheet['D13'].border = top_bold_border
  sheet['E13'].border = top_bold_border
  sheet['F13'].border = top_bold_border

  sheet['G13'].border = top_right_bold_border
  sheet['G14'].border = right_bold_border
  sheet['G15'].border = right_bold_border
  sheet['G16'].border = right_bold_border
  sheet['G17'].border = right_bold_border
  sheet['G18'].border = right_bold_border
  sheet['G19'].border = right_bold_border
  sheet['G20'].border = right_bold_border
  sheet['G21'].border = right_bold_border
  sheet['G22'].border = right_bold_border
  sheet['G23'].border = right_bold_border
  sheet['G24'].border = right_bold_border

  sheet['G25'].border = bottom_right_bold_border
  sheet['F25'].border = bottom_bold_border
  sheet['E25'].border = bottom_bold_border
  sheet['D25'].border = bottom_bold_border
  sheet['C25'].border = bottom_bold_border

  sheet['B25'].border = bottom_left_bold_border
  sheet['B24'].border = left_bold_border
  sheet['B23'].border = left_bold_border
  sheet['B22'].border = left_bold_border
  sheet['B21'].border = left_bold_border
  sheet['B20'].border = left_bold_border
  sheet['B19'].border = left_bold_border
  sheet['B18'].border = left_bold_border
  sheet['B17'].border = left_bold_border
  sheet['B16'].border = left_bold_border
  sheet['B15'].border = left_bold_border
  sheet['B14'].border = left_bold_border


  sheet['B28'].border = top_left_bold_border
  sheet['C28'].border = top_bold_border
  sheet['D28'].border = top_bold_border
  sheet['E28'].border = top_bold_border
  sheet['F28'].border = top_bold_border

  sheet['G28'].border = top_right_bold_border
  sheet['G29'].border = right_bold_border
  sheet['G30'].border = right_bold_border
  sheet['G31'].border = right_bold_border
  sheet['G32'].border = right_bold_border
  sheet['G33'].border = right_bold_border
  sheet['G34'].border = right_bold_border
  sheet['G35'].border = right_bold_border
  sheet['G36'].border = right_bold_border
  sheet['G37'].border = right_bold_border
  sheet['G38'].border = right_bold_border
  sheet['G39'].border = right_bold_border

  sheet['G40'].border = bottom_right_bold_border
  sheet['F40'].border = bottom_bold_border
  sheet['E40'].border = bottom_bold_border
  sheet['D40'].border = bottom_bold_border
  sheet['C40'].border = bottom_bold_border

  sheet['B40'].border = bottom_left_bold_border
  sheet['B39'].border = left_bold_border
  sheet['B38'].border = left_bold_border
  sheet['B37'].border = left_bold_border
  sheet['B36'].border = left_bold_border
  sheet['B35'].border = left_bold_border
  sheet['B34'].border = left_bold_border
  sheet['B33'].border = left_bold_border
  sheet['B32'].border = left_bold_border
  sheet['B31'].border = left_bold_border
  sheet['B30'].border = left_bold_border
  sheet['B29'].border = left_bold_border

  sheet['B12'] = '나. 전경사진'
  sheet['B27'] = '다. 위치도'

  frontViewPath = category.frontView.url[1:]
  locationMapPath = category.locationMap.url[1:]

  frontView = IMG.open(frontViewPath)
  frontWidth,frontHeight = frontView.size
  if (frontWidth < frontHeight):
    frontNewWidth = int((frontWidth/frontHeight) * baseHeight)
    frontViewImage = openpyxl.drawing.image.Image(frontViewPath)
    frontViewImage.width = frontNewWidth
    frontViewImage.height = baseHeight
    sheet.add_image(frontViewImage,"C14")
  else:
    frontNewHeight = int((frontHeight/frontWidth) * baseWidth)
    if (frontNewHeight > 260):
      frontNewWidth = int((frontWidth/frontHeight) * baseHeight)
      frontViewImage = openpyxl.drawing.image.Image(frontViewPath)
      frontViewImage.width = frontNewWidth
      frontViewImage.height = baseHeight
      
    else:
      frontViewImage = openpyxl.drawing.image.Image(frontViewPath)
      frontViewImage.width = baseWidth
      frontViewImage.height = frontNewHeight
    
    sheet.add_image(frontViewImage,"C14")

  
  locationMap = IMG.open(locationMapPath)
  locationWidth,locationHeight = locationMap.size

  if (locationWidth < locationHeight):
    locationNewWidth = int((locationWidth/locationHeight) * baseHeight)
    locationMapImage = openpyxl.drawing.image.Image(locationMapPath)
    locationMapImage.width = locationNewWidth
    locationMapImage.height = baseHeight
    sheet.add_image(locationMapImage,"C29")
  else:
    locationNewHeight = int((locationHeight/locationWidth) * baseWidth)
    if (locationNewHeight > 260):
      locationNewWidth = int((locationWidth/locationHeight) * baseHeight)
      locationMapImage = openpyxl.drawing.image.Image(locationMapPath)
      locationMapImage.width = locationNewWidth
      locationMapImage.height = baseHeight
      sheet.add_image(locationMapImage,"C29")
    else:
      locationMapImage = openpyxl.drawing.image.Image(locationMapPath)
      locationMapImage.width = baseWidth
      locationMapImage.height = locationNewHeight
      sheet.add_image(locationMapImage,"C29")

  sheet.sheet_view.view = "pageBreakPreview"
  for row in sheet.rows:
      for cell in row:
          cell.alignment = Alignment(horizontal="center", vertical="center")
  for column in range(66,72):
    sheet.column_dimensions[chr(column)].bestFit = True


  return wb


def looks(wb,pk):
  baseWidth = 215
  baseHeight = 162
  flatBaseWidth = 150
  flatBaseHeight = 150
  imgCell = 2
  infoCell = 10
  cellB = chr(66)
  cellC = chr(67)
  cellD = chr(68)
  sheet = wb.create_sheet("외관조사사진", 1)
  sheet = wb['외관조사사진']
  category = Category.objects.get(pk=pk)
  cracks = Crack.objects.filter(category__facilityName__icontains=category.facilityName)
  sheet.column_dimensions["A"].width = 1
  sheet.column_dimensions["B"].width = 27
  sheet.column_dimensions["C"].width = 1
  sheet.column_dimensions["D"].width = 27
  sheet.column_dimensions["E"].width = 1
  sheet.column_dimensions["F"].width = 27
  sheet.column_dimensions["G"].width = 1
  sheet.column_dimensions["H"].width = 27

  for crack in cracks:
    crackObj = CrackObj.objects.filter(parent=crack.id)
    numbering = crackObj.count()
    if numbering < 3:
      numbering = 0
    else:
      numbering = numbering-2
    crackObj = crackObj[numbering:]
    if crackObj.count() == 2:
      for crackObj in crackObj:
        path = crackObj.image.url[1:]
        flatPath = crackObj.flatting_image.url[1:]
        originImg = IMG.open(path) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        flatImg = IMG.open(flatPath) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        
        imgWidth, imgHeight = originImg.size
        flatImgWidth, flatImgHeight = flatImg.size

        if (imgWidth < imgHeight):
          imgNewWidth = int((imgWidth/imgHeight) * baseHeight)
          img = openpyxl.drawing.image.Image(path)
          img.width = imgNewWidth
          img.height = baseHeight
        else:
          imgNewHeight = int((imgHeight/imgWidth) * baseWidth)
          img = openpyxl.drawing.image.Image(path)
          img.width = baseWidth
          img.height = imgNewHeight

        if (flatImgWidth < flatImgHeight):
          flatImgNewWidth = int((flatImgWidth/flatImgHeight) * flatBaseHeight)
          flatImg = openpyxl.drawing.image.Image(flatPath)
          flatImg.width = flatImgNewWidth
          flatImg.height = flatBaseHeight
        else:
          flatImgNewHeight = int((flatImgHeight/flatImgWidth) * flatBaseWidth)
          flatImg = openpyxl.drawing.image.Image(flatPath)
          flatImg.width = flatBaseWidth
          flatImg.height = flatImgNewHeight

        sheet.add_image(img,cellB + str(imgCell))
        sheet.add_image(flatImg, cellD + str(imgCell))
        
        sheet[cellB+str(infoCell)] = '사진번호: ' + str(crackObj.id)
        sheet[cellB+str(infoCell+1)] = '위치: ' + str(crack.floor) + ' ' + str(crack.location)
        sheet[cellB+str(infoCell+2)] = '손상종류: ' + str(crack.crackType)
        sheet[cellD+str(infoCell+2)] = '손상규모: ' + str(crack.crackSize)
        sheet[cellB+str(infoCell+3)] = '발생원인: ' + str(crack.cause)
        sheet[cellD+str(infoCell+3)] = '적출년도: ' + str(crack.date)
        
        sheet[cellB+str(imgCell)].border = top_left_bold_border
        sheet[cellC+str(imgCell)].border = top_bold_border
        sheet[cellD+str(imgCell)].border = top_right_bold_border
        sheet[cellD+str(imgCell+1)].border = right_bold_border
        sheet[cellD+str(imgCell+2)].border = right_bold_border
        sheet[cellD+str(imgCell+3)].border = right_bold_border
        sheet[cellD+str(imgCell+4)].border = right_bold_border
        sheet[cellD+str(imgCell+5)].border = right_bold_border
        sheet[cellD+str(imgCell+6)].border = right_bold_border
        sheet[cellD+str(imgCell+7)].border = right_bold_border
        sheet[cellD+str(imgCell+8)].border = right_bold_border
        sheet[cellD+str(imgCell+9)].border = right_bold_border
        sheet[cellD+str(imgCell+10)].border = right_bold_border
        sheet[cellB+str(infoCell-7)].border = left_bold_border
        sheet[cellB+str(infoCell-6)].border = left_bold_border
        sheet[cellB+str(infoCell-5)].border = left_bold_border
        sheet[cellB+str(infoCell-4)].border = left_bold_border
        sheet[cellB+str(infoCell-3)].border = left_bold_border
        sheet[cellB+str(infoCell-2)].border = left_bold_border
        sheet[cellB+str(infoCell-1)].border = left_bold_border
        sheet[cellB+str(infoCell)].border = left_bold_border
        sheet[cellB+str(infoCell+1)].border = left_bold_border
        sheet[cellB+str(infoCell+2)].border = left_bold_border
        sheet[cellB+str(infoCell+3)].border = bottom_left_bold_border
        sheet[cellC+str(infoCell+3)].border = bottom_bold_border
        sheet[cellD+str(infoCell+3)].border = bottom_right_bold_border

        cellB = ord(cellB) + 4
        cellB = chr(cellB)
        cellC = ord(cellC) + 4
        cellC = chr(cellC)
        cellD = ord(cellD) + 4
        cellD = chr(cellD)
     
        if ord(cellB) > 72:
          cellB = chr(66)
          cellC = chr(67)
          infoCell += 7
          imgCell +=7
        if ord(cellD) > 72:
          cellD = chr(68)
          cellC = chr(67)
          imgCell += 7
          infoCell += 7
        sheet.sheet_view.view = "pageBreakPreview"
    else:
      cellB = chr(66)
      cellC = chr(67)
      cellD = chr(68)
      for crackObj in crackObj:
        path = crackObj.image.url[1:]
        flatPath = crackObj.flatting_image.url[1:]
        originImg = IMG.open(path) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        flatImg = IMG.open(flatPath) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        
        imgWidth, imgHeight = originImg.size
        flatImgWidth, flatImgHeight = flatImg.size

        if (imgWidth < imgHeight):
          imgNewWidth = int((imgWidth/imgHeight) * baseHeight)
          img = openpyxl.drawing.image.Image(path)
          img.width = imgNewWidth
          img.height = baseHeight
        
        else:
          imgNewHeight = int((imgHeight/imgWidth) * baseWidth)
          img = openpyxl.drawing.image.Image(path)
          img.width = baseWidth
          img.height = imgNewHeight
          

        if (flatImgWidth < flatImgHeight):
          flatImgNewWidth = int((flatImgWidth/flatImgHeight) * flatBaseHeight)
          flatImg = openpyxl.drawing.image.Image(flatPath)
          flatImg.width = flatImgNewWidth
          flatImg.height = flatBaseHeight
        else:
          flatImgNewHeight = int((flatImgHeight/flatImgWidth) * flatBaseWidth)
          flatImg = openpyxl.drawing.image.Image(flatPath)
          flatImg.width = flatBaseWidth
          flatImg.height = flatImgNewHeight

        sheet.add_image(img,cellB + str(imgCell))
        sheet.add_image(flatImg, cellD + str(imgCell))
        
        sheet[cellB+str(infoCell)] = '사진번호: ' + str(crackObj.id)
        sheet[cellB+str(infoCell+1)] = '위치: ' + str(crack.floor) + ' ' + str(crack.location)
        sheet[cellB+str(infoCell+2)] = '손상종류: ' + str(crack.crackType)
        sheet[cellD+str(infoCell+2)] = '손상규모: ' + str(crack.crackSize)
        sheet[cellB+str(infoCell+3)] = '발생원인: ' + str(crack.cause)
        sheet[cellD+str(infoCell+3)] = '적출년도: ' + str(crack.date)

        sheet[cellB+str(imgCell)].border = top_left_bold_border
        sheet[cellC+str(imgCell)].border = top_bold_border
        sheet[cellD+str(imgCell)].border = top_right_bold_border
        sheet[cellD+str(imgCell+1)].border = right_bold_border
        sheet[cellD+str(imgCell+2)].border = right_bold_border
        sheet[cellD+str(imgCell+3)].border = right_bold_border
        sheet[cellD+str(imgCell+4)].border = right_bold_border
        sheet[cellD+str(imgCell+5)].border = right_bold_border
        sheet[cellD+str(imgCell+6)].border = right_bold_border
        sheet[cellD+str(imgCell+7)].border = right_bold_border
        sheet[cellD+str(imgCell+8)].border = right_bold_border
        sheet[cellD+str(imgCell+9)].border = right_bold_border
        sheet[cellD+str(imgCell+10)].border = right_bold_border
        sheet[cellB+str(infoCell-7)].border = left_bold_border
        sheet[cellB+str(infoCell-6)].border = left_bold_border
        sheet[cellB+str(infoCell-5)].border = left_bold_border
        sheet[cellB+str(infoCell-4)].border = left_bold_border
        sheet[cellB+str(infoCell-3)].border = left_bold_border
        sheet[cellB+str(infoCell-2)].border = left_bold_border
        sheet[cellB+str(infoCell-1)].border = left_bold_border
        sheet[cellB+str(infoCell)].border = left_bold_border
        sheet[cellB+str(infoCell+1)].border = left_bold_border
        sheet[cellB+str(infoCell+2)].border = left_bold_border
        sheet[cellB+str(infoCell+3)].border = bottom_left_bold_border
        sheet[cellC+str(infoCell+3)].border = bottom_bold_border
        sheet[cellD+str(infoCell+3)].border = bottom_right_bold_border

        sheet.sheet_view.view = "pageBreakPreview"
  return wb
